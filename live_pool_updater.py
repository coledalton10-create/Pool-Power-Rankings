#!/usr/bin/env python3
"""
Live NCAA tournament pool updater using only The Odds API.

Data sources:
- /odds for remaining-game market probabilities
- NCAA scoreboard API for live scores, completed status, and winners

Outputs:
- live_pool_odds.csv
- live_pool_odds.xlsx
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import hashlib
import json
import math
import os
import shutil
import sys
import unicodedata
import urllib.parse
import urllib.request
import urllib.error

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from simulate_pool_portfolio import (
    FINAL_FOUR_PAIRINGS,
    TeamRecord,
    build_entries_matrix,
    load_first_four_games,
    load_model_inputs,
    normalize_text,
    read_pool_entries,
    resolve_pool_entries_path,
)


TOURNAMENT_FILE = "monte_carlo_tournament_model_2026_mens_fixed_torvik_barthag_zscores.xlsx"
OUTPUT_CSV = "live_pool_odds.csv"
OUTPUT_PREVIOUS_CSV = "live_pool_odds_previous.csv"
OUTPUT_XLSX = "live_pool_odds.xlsx"
OUTPUT_DASHBOARD_JSON = "live_pool_dashboard.json"
SNAPSHOT_STATE_FILE = "live_pool_snapshot_state.json"
RESOLVED_WINNERS_FILE = "resolved_winners.json"
DEFAULT_SIMULATIONS = 10_000
SITE_OUTPUT_DIR = "pages_site"

LIVE_SCORE_STATES = {"in_progress", "live", "halftime", "overtime"}
FINAL_SCORE_STATES = {"completed", "final"}
ROUND_SORT_PRIORITY = {68: 0, 64: 1, 32: 2, 16: 3, 8: 4, 4: 5, 2: 6}


@dataclass(frozen=True)
class BracketGame:
    node_id: str
    round_code: int
    region: str
    slot: int
    source_a: str
    source_b: str
    is_play_in: bool = False


@dataclass
class LiveNodeState:
    game: BracketGame
    fixed_winner: Optional[str] = None
    score_status: Optional[str] = None
    score_event_id: Optional[str] = None
    odds_event_id: Optional[str] = None
    market_probs: Optional[Dict[str, float]] = None
    matched_participants: Optional[Tuple[str, str]] = None


EXPLICIT_TEAM_ALIASES = {
    "uni": "Northern Iowa",
    "st johns ny": "St. John's",
    "st john's ny": "St. John's",
    "queens nc": "Queens",
    "prairie view": "Prairie View A&M",
    "byu cougars": "BYU",
    "hawaii": "Hawaiʻi",
    "hawaiʻi": "Hawaiʻi",
    "howard bison": "Howard",
    "michigan wolverines": "Michigan",
    "north carolina tar heels": "North Carolina",
    "saint marys gaels": "Saint Mary's",
    "saint mary's gaels": "Saint Mary's",
    "saint marys ca": "Saint Mary's",
    "saint mary's ca": "Saint Mary's",
    "south fla": "South Florida",
    "south fla.": "South Florida",
    "south florida": "South Florida",
    "texas am aggies": "Texas A&M",
    "texas a&m aggies": "Texas A&M",
    "usf": "South Florida",
    "vcu rams": "VCU",
    "ncsu": "NC State",
    "nc state": "NC State",
    "north carolina state": "NC State",
    "texas": "Texas",
    "texas longhorns": "Texas",
    "nc state texas winner": "NC State/Texas Winner",
    "nc state/texas winner": "NC State/Texas Winner",
}


def normalize_team_key(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return normalize_text(text)


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip("'").strip('"'))


def require_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise ValueError(f"Missing required environment variable: {name}")
    return value


def redact_api_key(value: str) -> str:
    if not value:
        return value
    if len(value) <= 8:
        return "<redacted>"
    return f"{value[:4]}...{value[-4:]}"


def redact_url(url: str, api_key: str) -> str:
    return url.replace(api_key, redact_api_key(api_key))


def fetch_json_with_headers(url: str, timeout: int = 30) -> Tuple[object, Dict[str, str]]:
    request = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        payload = response.read().decode("utf-8")
        headers = {key.lower(): value for key, value in response.headers.items()}
    return json.loads(payload), headers


def load_resolved_winners(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    return {str(key): str(value) for key, value in data.items()}


def save_resolved_winners(path: Path, winners: Dict[str, str]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        json.dump(winners, handle, indent=2, sort_keys=True)


def load_snapshot_state(path: Path) -> Dict[str, object]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        loaded = json.load(handle)
    return loaded if isinstance(loaded, dict) else {}


def save_snapshot_state(path: Path, state: Dict[str, object]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        json.dump(state, handle, indent=2, sort_keys=True)


def build_snapshot_key(resolved_winners: Dict[str, str]) -> str:
    payload = json.dumps(sorted(resolved_winners.items()), separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def get_scoreboard_base_date() -> datetime.date:
    base_date_raw = os.getenv("NCAA_SCOREBOARD_BASE_DATE", "").strip()
    if base_date_raw:
        return datetime.strptime(base_date_raw, "%Y-%m-%d").date()
    return datetime.now(timezone.utc).date()


def get_scoreboard_days() -> int:
    raw_days = os.getenv("NCAA_SCOREBOARD_DAYS", "2").strip() or "2"
    try:
        return max(1, int(raw_days))
    except ValueError:
        return 2


def get_simulation_count() -> int:
    raw_value = os.getenv("POOL_SIMULATIONS", str(DEFAULT_SIMULATIONS)).strip() or str(DEFAULT_SIMULATIONS)
    try:
        return max(1, int(raw_value))
    except ValueError:
        return DEFAULT_SIMULATIONS


def fetch_json_with_headers_verbose(
    url: str,
    *,
    api_key: str,
    label: str,
    timeout: int = 30,
) -> Tuple[object, Dict[str, str]]:
    redacted = redact_url(url, api_key)
    if label == "scores":
        print(f"Scores URL: {redacted}")
    if label == "odds":
        print(f"Odds URL: {redacted}")
    request = urllib.request.Request(url, headers={"Accept": "application/json"})
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = response.read().decode("utf-8")
            headers = {key.lower(): value for key, value in response.headers.items()}
        return json.loads(payload), headers
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        print(f"{label.capitalize()} request failed: {redacted}", file=sys.stderr)
        print(body, file=sys.stderr)
        raise


def append_query(url: str, params: Dict[str, str]) -> str:
    return f"{url}?{urllib.parse.urlencode(params)}"


def build_bracket_template(
    workbook_path: Path,
    resolved_winners: Optional[Dict[str, str]] = None,
) -> Dict[str, BracketGame]:
    first_four = load_first_four_games(workbook_path)
    first_round_df = pd.read_excel(workbook_path, sheet_name="first_round_matchups", engine="openpyxl")
    first_round_df.columns = [str(column).strip() for column in first_round_df.columns]
    first_round_df = first_round_df.dropna(how="all")
    first_round_df = first_round_df[first_round_df["game_id"].notna()].copy()
    first_round_df.loc[first_round_df["game_id"] == "R64_W_05", "team2_id"] = "FF_M_11"
    first_round_df.loc[first_round_df["game_id"] == "R64_S_01", "team2_id"] = "FF_M_16"
    first_round_df.loc[first_round_df["game_id"] == "R64_M_01", "team2_id"] = "FF_S_16"
    first_round_df.loc[first_round_df["game_id"] == "R64_M_05", "team2_id"] = "FF_W_11"

    nodes: Dict[str, BracketGame] = {}
    resolved_winners = resolved_winners or {}

    for placeholder_id, (team1_id, team2_id) in first_four.items():
        region_code = placeholder_id.split("_")[1]
        region_name = {"E": "East", "W": "West", "S": "South", "M": "Midwest"}.get(region_code, region_code)
        nodes[placeholder_id] = BracketGame(
            node_id=placeholder_id,
            round_code=68,
            region=region_name,
            slot=int(placeholder_id.split("_")[-1]),
            source_a=team1_id,
            source_b=team2_id,
            is_play_in=True,
        )

    region_round64: Dict[str, List[BracketGame]] = {}
    for row in first_round_df.itertuples(index=False):
        source_a = str(row.team1_id).strip()
        source_b = str(row.team2_id).strip()
        if source_a.startswith("FF_") and source_a in resolved_winners:
            source_a = resolved_winners[source_a]
        if source_b.startswith("FF_") and source_b in resolved_winners:
            source_b = resolved_winners[source_b]
        game = BracketGame(
            node_id=str(row.game_id).strip(),
            round_code=int(row.round_num),
            region=str(row.region).strip(),
            slot=int(row.slot),
            source_a=source_a,
            source_b=source_b,
        )
        nodes[game.node_id] = game
        region_round64.setdefault(game.region, []).append(game)

    for region, games in region_round64.items():
        current_round = sorted(games, key=lambda item: item.slot)
        for round_code in (32, 16, 8):
            next_round: List[BracketGame] = []
            for idx in range(0, len(current_round), 2):
                left = current_round[idx]
                right = current_round[idx + 1]
                slot = idx // 2 + 1
                node_id = f"R{round_code}_{region[:1].upper()}_{slot:02d}"
                game = BracketGame(
                    node_id=node_id,
                    round_code=round_code,
                    region=region,
                    slot=slot,
                    source_a=left.node_id,
                    source_b=right.node_id,
                )
                nodes[node_id] = game
                next_round.append(game)
            current_round = next_round

    semifinal_nodes: List[BracketGame] = []
    for slot, (region_a, region_b) in enumerate(FINAL_FOUR_PAIRINGS, start=1):
        node = BracketGame(
            node_id=f"R4_FF_{slot:02d}",
            round_code=4,
            region="Final Four",
            slot=slot,
            source_a=f"R8_{region_a[:1].upper()}_01",
            source_b=f"R8_{region_b[:1].upper()}_01",
        )
        nodes[node.node_id] = node
        semifinal_nodes.append(node)

    nodes["R2_CHAMP_01"] = BracketGame(
        node_id="R2_CHAMP_01",
        round_code=2,
        region="Championship",
        slot=1,
        source_a=semifinal_nodes[0].node_id,
        source_b=semifinal_nodes[1].node_id,
    )

    return nodes


def build_team_name_index(team_lookup: Dict[str, TeamRecord]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for team in team_lookup.values():
        variants = {
            team.team_name,
            team.team_id,
            team.team_name.replace("'", ""),
            team.team_name.replace("’", ""),
        }
        for variant in variants:
            normalized = normalize_team_key(variant)
            if normalized:
                mapping.setdefault(normalized, team.team_name)

    aliases = {
        "mich st": "Michigan State",
        "mich state": "Michigan State",
        "michigan st": "Michigan State",
        "saint marys": "Saint Mary's",
        "saint mary's": "Saint Mary's",
        "st marys": "Saint Mary's",
        "st mary's": "Saint Mary's",
        "st johns": "St. John's",
        "saint johns": "St. John's",
        "usf": "South Florida",
        "south florida": "South Florida",
        "unc": "North Carolina",
        "miami fl": "Miami (FL)",
        "miami florida": "Miami (FL)",
        "miami oh": "Miami (OH)",
        "miami ohio": "Miami (OH)",
        "uconn": "UConn",
        "connecticut": "UConn",
    }
    existing_names = {team.team_name for team in team_lookup.values()}
    for alias, target in aliases.items():
        if target in existing_names:
            mapping[normalize_team_key(alias)] = target
    for alias, target in EXPLICIT_TEAM_ALIASES.items():
        if target in existing_names:
            mapping[normalize_team_key(alias)] = target
    return mapping


def build_team_id_index(team_lookup: Dict[str, TeamRecord]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for team in team_lookup.values():
        mapping[str(team.team_id).strip()] = team.team_name
    return mapping


def match_external_team(raw_value: object, team_name_index: Dict[str, str]) -> Optional[str]:
    if raw_value is None:
        return None
    normalized = normalize_team_key(raw_value)
    if not normalized:
        return None
    if normalized in EXPLICIT_TEAM_ALIASES:
        return EXPLICIT_TEAM_ALIASES[normalized]
    return team_name_index.get(normalized)


def canonicalize_name(
    raw_value: object,
    team_name_index: Dict[str, str],
    team_id_index: Optional[Dict[str, str]] = None,
) -> Optional[str]:
    if raw_value is None:
        return None
    raw_text = str(raw_value).strip()
    if team_id_index and raw_text in team_id_index:
        return team_id_index[raw_text]
    normalized = normalize_team_key(raw_text)
    if not normalized:
        return None
    if normalized in EXPLICIT_TEAM_ALIASES:
        return EXPLICIT_TEAM_ALIASES[normalized]
    return team_name_index.get(normalized)


def strip_team_nickname(raw_value: object) -> str:
    text = str(raw_value or "").strip()
    if not text:
        return ""
    normalized = normalize_team_key(text)
    tokens = normalized.split()
    if len(tokens) <= 1:
        return normalized
    removable_suffixes = {
        "aggies",
        "bears",
        "bison",
        "boilermakers",
        "broncos",
        "bruins",
        "buckeyes",
        "cardinals",
        "cougars",
        "eagles",
        "golden",
        "wolfpack",
        "longhorns",
        "lions",
        "lobo",
        "lobos",
        "owls",
        "panthers",
        "pirates",
        "rams",
        "rebels",
        "spiders",
        "spartans",
        "terriers",
        "tigers",
        "trojans",
        "huskies",
        "bulldogs",
        "wildcats",
        "hurricanes",
        "redstorm",
        "red",
        "storm",
        "gaels",
        "gators",
        "blue",
        "devils",
        "tar",
        "heels",
        "demons",
        "hawks",
        "knights",
        "orange",
        "raiders",
    }
    while len(tokens) > 1 and tokens[-1] in removable_suffixes:
        tokens.pop()
    return " ".join(tokens)


def canonical_team_pair_key(
    team_a: object,
    team_b: object,
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> Optional[frozenset]:
    candidates = []
    for raw in (team_a, team_b):
        canonical = canonicalize_name(raw, team_name_index, team_id_index)
        if not canonical:
            stripped = strip_team_nickname(raw)
            canonical = canonicalize_name(stripped, team_name_index, team_id_index)
        if not canonical:
            return None
        candidates.append(canonical)
    return frozenset(candidates)


def tournament_window() -> Tuple[str, str]:
    now = datetime.combine(get_scoreboard_base_date(), datetime.min.time(), tzinfo=timezone.utc)
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end = start.replace(hour=23, minute=59, second=59)
    return start.isoformat().replace("+00:00", "Z"), end.isoformat().replace("+00:00", "Z")


def american_to_decimal(odds: float) -> float:
    if odds > 0:
        return 1.0 + (odds / 100.0)
    return 1.0 + (100.0 / abs(odds))


def _safe_float(value: object) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def outcome_price_to_decimal(price: object) -> Optional[float]:
    value = _safe_float(price)
    if value is None:
        return None
    if value > 1000 or value < -100:
        return american_to_decimal(value)
    if value > 1:
        return value
    return None


def _iter_tournament_scoreboard_dates() -> List[datetime]:
    today = get_scoreboard_base_date()
    days = get_scoreboard_days()
    return [
        datetime.combine(today - timedelta(days=offset), datetime.min.time(), tzinfo=timezone.utc)
        for offset in range(days)
    ]


def normalize_ncaa_scoreboard_game(raw_item: dict) -> Optional[dict]:
    game = raw_item.get("game", raw_item)
    if not isinstance(game, dict):
        return None

    game_id = game.get("gameID") or game.get("game_id") or game.get("id")
    away = game.get("away", {}) if isinstance(game.get("away"), dict) else {}
    home = game.get("home", {}) if isinstance(game.get("home"), dict) else {}
    away_names = away.get("names", {}) if isinstance(away.get("names"), dict) else {}
    home_names = home.get("names", {}) if isinstance(home.get("names"), dict) else {}

    away_team = (
        away_names.get("short")
        or away_names.get("seo")
        or away_names.get("char6")
        or away.get("name")
    )
    home_team = (
        home_names.get("short")
        or home_names.get("seo")
        or home_names.get("char6")
        or home.get("name")
    )
    away_score = away.get("score")
    home_score = home.get("score")
    game_state = str(game.get("gameState") or game.get("game_state") or "").strip().lower()
    final_flag = bool(game.get("final"))
    live_flag = bool(game.get("live"))

    if not game_id or not away_team or not home_team:
        return None

    if final_flag or game_state in {"final", "official"}:
        status = "completed"
        completed = True
    elif live_flag or game_state in {"live", "in progress", "halftime"}:
        status = "in_progress"
        completed = False
    else:
        status = game_state or "scheduled"
        completed = False

    return {
        "id": str(game_id),
        "completed": completed,
        "status": status,
        "away_team": away_team,
        "home_team": home_team,
        "scores": [
            {"name": away_team, "score": away_score},
            {"name": home_team, "score": home_score},
        ],
    }


def load_ncaa_scoreboard_games() -> Tuple[List[dict], Optional[str]]:
    normalized_games: List[dict] = []
    for dt in _iter_tournament_scoreboard_dates():
        url = (
            "https://ncaa-api.henrygd.me/scoreboard/"
            f"basketball-men/d1/{dt:%Y/%m/%d}/all-conf"
        )
        print(f"NCAA scores URL: {url}")
        payload, _headers = fetch_json_with_headers(url)

        games = payload.get("games", []) if isinstance(payload, dict) else []
        for raw_item in games:
            normalized = normalize_ncaa_scoreboard_game(raw_item)
            if normalized is not None:
                normalized_games.append(normalized)

    deduped: Dict[str, dict] = {}
    for game in normalized_games:
        deduped[game["id"]] = game

    return list(deduped.values()), None


def load_the_odds_api_odds(score_events: List[dict]) -> Tuple[List[dict], Optional[str]]:
    api_key = require_env("THE_ODDS_API_KEY")
    base = "https://api.the-odds-api.com/v4"
    sport_key = "basketball_ncaab"
    commence_from, commence_to = tournament_window()

    params = {
        "apiKey": api_key,
        "regions": "us",
        "markets": "h2h",
        "oddsFormat": "decimal",
        "commenceTimeFrom": commence_from,
        "commenceTimeTo": commence_to,
    }

    odds_url = append_query(f"{base}/sports/{sport_key}/odds", params)
    odds_payload, odds_headers = fetch_json_with_headers_verbose(
        odds_url,
        api_key=api_key,
        label="odds",
    )
    if not isinstance(odds_payload, list):
        raise ValueError("Expected The Odds API /odds endpoint to return a list.")

    credits_remaining = odds_headers.get("x-requests-remaining") or odds_headers.get("x-requests-used")
    return odds_payload, credits_remaining


def build_market_probabilities(
    odds_events: List[dict],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> Tuple[Dict[str, Dict[str, float]], Dict[frozenset, Dict[str, float]], int, int]:
    by_event_id: Dict[str, Dict[str, float]] = {}
    by_pair: Dict[frozenset, List[Dict[str, float]]] = {}
    before_count = len(odds_events)
    matched_count = 0

    for event in odds_events:
        event_id = str(event.get("id") or "")
        pair_key = canonical_team_pair_key(
            event.get("home_team"),
            event.get("away_team"),
            team_name_index,
            team_id_index,
        )
        if not pair_key:
            pair_key = canonical_team_pair_key(
                strip_team_nickname(event.get("home_team")),
                strip_team_nickname(event.get("away_team")),
                team_name_index,
                team_id_index,
            )
        if not event_id or not pair_key or len(pair_key) != 2:
            continue
        matched_count += 1
        for bookmaker in event.get("bookmakers", []):
            for market in bookmaker.get("markets", []):
                if market.get("key") != "h2h":
                    continue
                prices: Dict[str, float] = {}
                for outcome in market.get("outcomes", []):
                    team = canonicalize_name(outcome.get("name"), team_name_index, team_id_index)
                    if not team:
                        team = canonicalize_name(strip_team_nickname(outcome.get("name")), team_name_index, team_id_index)
                    decimal_price = outcome_price_to_decimal(outcome.get("price"))
                    if team and decimal_price:
                        prices[team] = decimal_price
                if len(prices) == 2:
                    implied = {team: 1.0 / price for team, price in prices.items()}
                    total = sum(implied.values())
                    normalized = {team: value / total for team, value in implied.items()}
                    by_event_id.setdefault(event_id, normalized)
                    by_pair.setdefault(pair_key, []).append(normalized)

    averaged_by_pair: Dict[frozenset, Dict[str, float]] = {}
    for pair_key, samples in by_pair.items():
        teams = list(pair_key)
        averaged = {team: float(np.mean([sample[team] for sample in samples])) for team in teams}
        total = sum(averaged.values())
        averaged_by_pair[pair_key] = {team: value / total for team, value in averaged.items()}

    return by_event_id, averaged_by_pair, before_count, matched_count


def extract_scores_participants(
    score_event: dict,
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> Tuple[Optional[str], Optional[str]]:
    scores = score_event.get("scores", [])
    if isinstance(scores, list) and len(scores) >= 2:
        teams = []
        for item in scores[:2]:
            team = canonicalize_name(item.get("name"), team_name_index, team_id_index)
            teams.append(team)
        return teams[0], teams[1]

    home_team = canonicalize_name(score_event.get("home_team"), team_name_index, team_id_index)
    away_team = canonicalize_name(score_event.get("away_team"), team_name_index, team_id_index)
    return away_team, home_team


def extract_score_winner(
    score_event: dict,
    participants: Tuple[Optional[str], Optional[str]],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> Optional[str]:
    scores = score_event.get("scores", [])
    if isinstance(scores, list) and len(scores) >= 2:
        resolved: List[Tuple[Optional[str], Optional[float]]] = []
        for item in scores[:2]:
            team = canonicalize_name(item.get("name"), team_name_index, team_id_index)
            resolved.append((team, _safe_float(item.get("score"))))
        if len(resolved) == 2 and resolved[0][0] and resolved[1][0]:
            if resolved[0][1] is not None and resolved[1][1] is not None and resolved[0][1] != resolved[1][1]:
                return resolved[0][0] if resolved[0][1] > resolved[1][1] else resolved[1][0]
    return None


def resolve_source(
    source: str,
    winners: Dict[str, str],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> Optional[str]:
    if source in winners:
        return winners[source]
    if source.startswith("R") or source.startswith("FF_"):
        return winners.get(source)
    return canonicalize_name(source, team_name_index, team_id_index)


def infer_participants(
    node: BracketGame,
    winners: Dict[str, str],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> Optional[Tuple[str, str]]:
    team_a = resolve_source(node.source_a, winners, team_name_index, team_id_index)
    team_b = resolve_source(node.source_b, winners, team_name_index, team_id_index)
    if team_a and team_b:
        return team_a, team_b
    return None


def build_live_state(
    nodes: Dict[str, BracketGame],
    score_events: List[dict],
    odds_by_event_id: Dict[str, Dict[str, float]],
    odds_by_pair: Dict[frozenset, Dict[str, float]],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
    preloaded_winners: Optional[Dict[str, str]] = None,
) -> Tuple[Dict[str, LiveNodeState], Dict[str, str], Dict[str, int]]:
    live_state = {node_id: LiveNodeState(game=node) for node_id, node in nodes.items()}
    winners: Dict[str, str] = dict(preloaded_winners or {})

    for node_id, winner in winners.items():
        if node_id in live_state:
            live_state[node_id].fixed_winner = winner

    normalized_scores: List[dict] = []
    for event in score_events:
        event_id = str(event.get("id") or "")
        participants = extract_scores_participants(event, team_name_index, team_id_index)
        if not participants[0] or not participants[1]:
            continue
        completed = bool(event.get("completed"))
        status = "completed" if completed else str(event.get("status") or "").strip().lower()
        winner = extract_score_winner(event, participants, team_name_index, team_id_index)
        normalized_scores.append(
            {
                "event_id": event_id,
                "participants": tuple(sorted((participants[0], participants[1]))),
                "ordered_participants": (participants[0], participants[1]),
                "status": status,
                "winner": winner,
            }
        )

    ordered_nodes = sorted(
        nodes.values(),
        key=lambda item: (ROUND_SORT_PRIORITY.get(item.round_code, 99), item.region, item.slot),
    )

    matched_score_events = 0
    for node in ordered_nodes:
        state = live_state[node.node_id]
        participants = infer_participants(node, winners, team_name_index, team_id_index)
        if not participants:
            continue

        state.matched_participants = participants
        pair_key = tuple(sorted(participants))
        matched_score = next((event for event in normalized_scores if event["participants"] == pair_key), None)

        if matched_score:
            matched_score_events += 1
            state.score_event_id = matched_score["event_id"]
            state.score_status = matched_score["status"]
            if matched_score["status"] in FINAL_SCORE_STATES and matched_score["winner"]:
                state.fixed_winner = matched_score["winner"]
                winners[node.node_id] = matched_score["winner"]
                continue

            if matched_score["event_id"] and matched_score["event_id"] in odds_by_event_id:
                state.odds_event_id = matched_score["event_id"]
                state.market_probs = odds_by_event_id[matched_score["event_id"]]

        if not state.market_probs:
            pair_market = odds_by_pair.get(frozenset(participants))
            if pair_market:
                state.market_probs = pair_market

    counts = {
        "completed_games": 0,
        "live_games": 0,
        "remaining_games_with_odds": 0,
        "remaining_games_fallback": 0,
    }
    scoring_rounds = {64, 32, 16, 8, 4, 2}
    for node in ordered_nodes:
        if node.round_code not in scoring_rounds:
            continue
        state = live_state[node.node_id]
        if state.fixed_winner:
            counts["completed_games"] += 1
        elif state.score_status in LIVE_SCORE_STATES:
            counts["live_games"] += 1
        elif state.market_probs:
            counts["remaining_games_with_odds"] += 1
        else:
            counts["remaining_games_fallback"] += 1

    counts["matched_score_events"] = matched_score_events
    counts["scores_events_returned"] = len(normalized_scores)
    return live_state, winners, counts


def logistic_win_probability(rating_a: float, rating_b: float, logistic_k: float) -> float:
    exponent = -logistic_k * (rating_a - rating_b)
    exponent = max(min(exponent, 700), -700)
    return 1.0 / (1.0 + math.exp(exponent))


def simulate_remaining_bracket(
    nodes: Dict[str, BracketGame],
    live_state: Dict[str, LiveNodeState],
    fixed_winners: Dict[str, str],
    team_lookup_by_name: Dict[str, TeamRecord],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
    round_multipliers: Dict[int, int],
    logistic_k: float,
    rng: np.random.Generator,
) -> Dict[str, int]:
    winners = dict(fixed_winners)
    additional_scores: Dict[str, int] = {}

    node_order = sorted(
        nodes.values(),
        key=lambda item: (ROUND_SORT_PRIORITY.get(item.round_code, 99), item.region, item.slot),
    )

    for node in node_order:
        state = live_state[node.node_id]
        if state.fixed_winner:
            continue

        team_a = resolve_source(node.source_a, winners, team_name_index, team_id_index)
        team_b = resolve_source(node.source_b, winners, team_name_index, team_id_index)
        if not team_a or not team_b:
            raise ValueError(f"Could not resolve participants for {node.node_id}.")

        if state.market_probs and team_a in state.market_probs and team_b in state.market_probs:
            probability_a = state.market_probs[team_a]
        else:
            probability_a = logistic_win_probability(
                team_lookup_by_name[team_a].rating,
                team_lookup_by_name[team_b].rating,
                logistic_k,
            )

        winner = team_a if rng.random() < probability_a else team_b
        winners[node.node_id] = winner
        if node.round_code in round_multipliers:
            additional_scores[winner] = additional_scores.get(winner, 0) + (
                team_lookup_by_name[winner].base_points * round_multipliers[node.round_code]
            )

    return additional_scores


def find_unresolved_bracket_names(
    nodes: Dict[str, BracketGame],
    fixed_winners: Dict[str, str],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> List[str]:
    unresolved = set()
    for node in sorted(nodes.values(), key=lambda item: (0 if item.round_code == 68 else item.round_code, item.region, item.slot)):
        for source in (node.source_a, node.source_b):
            if source in fixed_winners:
                continue
            if source.startswith("R") or source.startswith("FF_"):
                continue
            if canonicalize_name(source, team_name_index, team_id_index) is None:
                unresolved.add(str(source))
    return sorted(unresolved)


def validate_first_four_placeholders(
    nodes: Dict[str, BracketGame],
    fixed_winners: Dict[str, str],
    team_name_index: Dict[str, str],
    team_id_index: Dict[str, str],
) -> Tuple[str, List[str]]:
    placeholders = sorted(
        [node for node in nodes.values() if node.is_play_in],
        key=lambda item: (item.region, item.slot, item.node_id),
    )
    lines = ["First Four validation", "====================="]
    unresolved_placeholders: List[str] = []
    for node in placeholders:
        team_a = canonicalize_name(node.source_a, team_name_index, team_id_index)
        team_b = canonicalize_name(node.source_b, team_name_index, team_id_index)
        if node.node_id in fixed_winners:
            lines.append(f" - {node.node_id}: resolved from live results -> {fixed_winners[node.node_id]}")
        elif team_a and team_b:
            lines.append(f" - {node.node_id}: unresolved live result, will simulate {team_a} vs {team_b}")
        else:
            available = [candidate for candidate in [team_a, team_b] if candidate]
            lines.append(
                f" - {node.node_id}: unresolved placeholder; expected {node.source_a} vs {node.source_b}; "
                f"available candidate teams: {', '.join(available) if available else 'none'}"
            )
            unresolved_placeholders.append(node.node_id)
    return "\n".join(lines), unresolved_placeholders


def build_zero_counts_diagnostic(
    odds_events: List[dict],
    score_events: List[dict],
    counts: Dict[str, int],
    team_name_index: Dict[str, str],
) -> str:
    if not (
        counts["completed_games"] == 0
        and counts["live_games"] == 0
        and counts["remaining_games_with_odds"] == 0
    ):
        return ""

    matched_odds_events = 0
    for event in odds_events:
        home = match_external_team(event.get("home_team"), team_name_index)
        away = match_external_team(event.get("away_team"), team_name_index)
        if home and away:
            matched_odds_events += 1

    matched_score_events = 0
    for event in score_events:
        away, home = extract_scores_participants(event, team_name_index, team_id_index)
        if away and home:
            matched_score_events += 1

    sample_score_names = []
    for event in score_events[:3]:
        scores = event.get("scores", [])
        names = [str(item.get("name")) for item in scores[:2] if item.get("name")]
        if not names:
            names = [str(event.get("away_team") or ""), str(event.get("home_team") or "")]
        sample_score_names.append(" vs ".join([name for name in names if name]))

    sample_odds_names = []
    for event in odds_events[:3]:
        names = [str(event.get("away_team") or ""), str(event.get("home_team") or "")]
        sample_odds_names.append(" vs ".join([name for name in names if name]))

    lines = [
        "",
        "Zero-count diagnostic",
        "=====================",
        f"NCAA scoreboard events returned: {len(score_events)}",
        f"/odds events returned: {len(odds_events)}",
        f"NCAA scoreboard events matched to bracket-team names: {matched_score_events}",
        f"/odds events matched to bracket-team names: {matched_odds_events}",
        "Sample NCAA scoreboard teams:",
    ]
    lines.extend([f" - {sample}" for sample in sample_score_names] or [" - None"])
    lines.append("Sample /odds teams:")
    lines.extend([f" - {sample}" for sample in sample_odds_names] or [" - None"])
    return "\n".join(lines)


def compute_current_state(
    nodes: Dict[str, BracketGame],
    live_state: Dict[str, LiveNodeState],
    team_lookup_by_name: Dict[str, TeamRecord],
    round_multipliers: Dict[int, int],
) -> Tuple[Dict[str, int], set[str]]:
    current_scores: Dict[str, int] = {}
    eliminated: set[str] = set()

    for node in sorted(nodes.values(), key=lambda item: (item.round_code, item.region, item.slot)):
        state = live_state[node.node_id]
        if not state.fixed_winner:
            continue
        if state.matched_participants:
            loser = state.matched_participants[0] if state.fixed_winner == state.matched_participants[1] else state.matched_participants[1]
            eliminated.add(loser)
        if node.round_code in round_multipliers:
            current_scores[state.fixed_winner] = current_scores.get(state.fixed_winner, 0) + (
                team_lookup_by_name[state.fixed_winner].base_points * round_multipliers[node.round_code]
            )

    return current_scores, eliminated


def calculate_finish_positions(scores: np.ndarray) -> np.ndarray:
    finishes = np.empty_like(scores, dtype=np.int16)
    for idx, score in enumerate(scores):
        finishes[idx] = 1 + int(np.sum(scores > score))
    return finishes


def calculate_finish_metrics(
    scores: np.ndarray,
    *,
    top_finish_cutoffs: Tuple[int, ...] = (1, 3, 4, 10),
) -> Tuple[np.ndarray, Dict[int, np.ndarray], np.ndarray]:
    """Return fair tied-finish metrics for one simulated score vector.

    For tied groups, finish value is the average occupied placement and top-k
    shares are split proportionally across the tied entrants.
    """
    entrant_count = len(scores)
    finish_values = np.zeros(entrant_count, dtype=float)
    top_shares = {cutoff: np.zeros(entrant_count, dtype=float) for cutoff in top_finish_cutoffs}
    last_place_share = np.zeros(entrant_count, dtype=float)

    ranked_indices = np.argsort(-scores, kind="stable")
    ranked_scores = scores[ranked_indices]
    start_idx = 0
    current_position = 1

    while start_idx < entrant_count:
        end_idx = start_idx + 1
        while end_idx < entrant_count and ranked_scores[end_idx] == ranked_scores[start_idx]:
            end_idx += 1

        tied_indices = ranked_indices[start_idx:end_idx]
        group_size = end_idx - start_idx
        group_start = current_position
        group_end = current_position + group_size - 1

        finish_values[tied_indices] = (group_start + group_end) / 2.0
        for cutoff in top_finish_cutoffs:
            overlap = max(0, min(group_end, cutoff) - group_start + 1)
            if overlap:
                top_shares[cutoff][tied_indices] = overlap / group_size
        if group_start <= entrant_count <= group_end:
            last_place_share[tied_indices] = 1.0 / group_size

        current_position = group_end + 1
        start_idx = end_idx

    return finish_values, top_shares, last_place_share


def build_validation_summary(counts: Dict[str, int], credits_remaining: Optional[str]) -> str:
    lines = [
        "Validation summary",
        "==================",
        f"Completed games: {counts['completed_games']}",
        f"Live games: {counts['live_games']}",
        f"Remaining games with odds: {counts['remaining_games_with_odds']}",
        f"Remaining games using fallback model: {counts['remaining_games_fallback']}",
        f"Credits remaining: {credits_remaining or 'unknown'}",
    ]
    return "\n".join(lines)


def summarize_results(
    entries_df: pd.DataFrame,
    current_scores: np.ndarray,
    sim_score_matrix: np.ndarray,
    finish_matrix: np.ndarray,
    win_share_matrix: np.ndarray,
    top3_share_matrix: np.ndarray,
    top4_share_matrix: np.ndarray,
    top10_share_matrix: np.ndarray,
    last_share_matrix: np.ndarray,
    eliminated_teams: set[str],
) -> pd.DataFrame:
    total_scores = sim_score_matrix + current_scores[np.newaxis, :]
    summary = pd.DataFrame(
        {
            "name": entries_df["name"].astype(str),
            "current_score": current_scores,
            "live_win_rate": win_share_matrix.mean(axis=0),
            "top3_rate": top3_share_matrix.mean(axis=0),
            "top4_rate": top4_share_matrix.mean(axis=0),
            "avg_finish": finish_matrix.mean(axis=0),
            "remaining_live_teams": entries_df["teams"].apply(
                lambda text: sum(team.strip() not in eliminated_teams for team in str(text).split(","))
            ),
            "p25_score": np.percentile(total_scores, 25, axis=0),
            "p50_score": np.percentile(total_scores, 50, axis=0),
            "p75_score": np.percentile(total_scores, 75, axis=0),
            "p90_score": np.percentile(total_scores, 90, axis=0),
            "p95_score": np.percentile(total_scores, 95, axis=0),
            "best_case_score": total_scores.max(axis=0),
            "worst_case_score": total_scores.min(axis=0),
            "teams": entries_df["teams"],
            "last_updated_at": datetime.now(timezone.utc).isoformat(),
        }
    )
    summary["current_rank"] = summary["current_score"].rank(method="min", ascending=False)
    summary["live_rank"] = summary["live_win_rate"].rank(method="min", ascending=False)
    summary["movement"] = summary["current_rank"] - summary["live_rank"]
    summary["top10_rate"] = top10_share_matrix.mean(axis=0)
    summary["last_rate"] = last_share_matrix.mean(axis=0)
    summary["first_place_tie_rate"] = ((win_share_matrix > 0) & (win_share_matrix < 1)).mean(axis=0)
    return summary


def describe_completed_games(
    nodes: Dict[str, BracketGame],
    live_state: Dict[str, LiveNodeState],
    previous_winners: Dict[str, str],
    current_winners: Dict[str, str],
) -> List[str]:
    round_labels = {
        68: "First Four",
        64: "Round of 64",
        32: "Round of 32",
        16: "Sweet 16",
        8: "Elite 8",
        4: "Final Four",
        2: "Championship",
    }
    descriptions: List[str] = []
    for node_id, winner in sorted(current_winners.items(), key=lambda item: (ROUND_SORT_PRIORITY.get(nodes[item[0]].round_code, 99), nodes[item[0]].region, nodes[item[0]].slot)):
        if previous_winners.get(node_id) == winner:
            continue
        state = live_state[node_id]
        participants = state.matched_participants or (None, None)
        loser = None
        if participants[0] and participants[1]:
            loser = participants[0] if winner == participants[1] else participants[1]
        round_label = round_labels.get(nodes[node_id].round_code, str(nodes[node_id].round_code))
        if loser:
            descriptions.append(f"{round_label}: {winner} over {loser}")
        else:
            descriptions.append(f"{round_label}: {winner}")
    return descriptions


def apply_trend_baseline(summary: pd.DataFrame, baseline_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    trended = summary.copy()
    trended["trend_rank_delta"] = 0.0
    trended["trend_score_delta"] = 0.0
    trended["trend_odds_delta"] = 0.0
    trended["trend_rank_direction"] = "flat"

    if baseline_df is None or baseline_df.empty:
        return trended

    baseline_rows = baseline_df.set_index("name")
    for idx, row in trended.iterrows():
        name = row["name"]
        if name not in baseline_rows.index:
            continue
        baseline_row = baseline_rows.loc[name]
        rank_delta = float(baseline_row["current_rank"]) - float(row["current_rank"])
        score_delta = float(row["current_score"]) - float(baseline_row["current_score"])
        odds_delta = float(row["live_win_rate"]) - float(baseline_row["live_win_rate"])
        trended.at[idx, "trend_rank_delta"] = rank_delta
        trended.at[idx, "trend_score_delta"] = score_delta
        trended.at[idx, "trend_odds_delta"] = odds_delta
        if rank_delta > 0:
            trended.at[idx, "trend_rank_direction"] = "up"
        elif rank_delta < 0:
            trended.at[idx, "trend_rank_direction"] = "down"

    return trended


def determine_current_round(nodes: Dict[str, BracketGame], live_state: Dict[str, LiveNodeState]) -> Optional[int]:
    scoring_rounds = [64, 32, 16, 8, 4, 2]
    for round_code in scoring_rounds:
        relevant_nodes = [node for node in nodes.values() if node.round_code == round_code]
        if any(live_state[node.node_id].fixed_winner is None for node in relevant_nodes):
            return round_code
    return None


def build_team_status_map(
    nodes: Dict[str, BracketGame],
    live_state: Dict[str, LiveNodeState],
) -> Tuple[Dict[str, Dict[str, object]], Optional[int]]:
    current_round = determine_current_round(nodes, live_state)
    status_map: Dict[str, Dict[str, object]] = {}

    if current_round is None:
        return status_map, None

    for node in nodes.values():
        state = live_state[node.node_id]
        participants = state.matched_participants or (None, None)
        if not participants[0] or not participants[1]:
            continue
        if node.round_code == current_round:
            for team_name in participants:
                status_map.setdefault(team_name, {"current_round_status": "not_played_current_round", "alive": True})
            if state.fixed_winner:
                loser = participants[0] if state.fixed_winner == participants[1] else participants[1]
                status_map[state.fixed_winner] = {"current_round_status": "won_current_round", "alive": True}
                status_map[loser] = {"current_round_status": "lost_current_round", "alive": False}

    for node in nodes.values():
        state = live_state[node.node_id]
        participants = state.matched_participants or (None, None)
        if not participants[0] or not participants[1] or not state.fixed_winner:
            continue
        loser = participants[0] if state.fixed_winner == participants[1] else participants[1]
        status_map.setdefault(state.fixed_winner, {"current_round_status": "advanced_prior", "alive": True})
        existing = status_map.get(loser, {})
        if existing.get("current_round_status") != "lost_current_round":
            status_map[loser] = {"current_round_status": existing.get("current_round_status", "eliminated_prior"), "alive": False}

    return status_map, current_round


def build_first_four_pick_resolution_map(
    first_four_lookup: Dict[str, Tuple[str, str]],
    team_lookup: Dict[str, TeamRecord],
    resolved_winners: Dict[str, str],
) -> Dict[str, str]:
    resolution_map: Dict[str, str] = {}

    for placeholder_id, (team1_id, team2_id) in first_four_lookup.items():
        winner = resolved_winners.get(placeholder_id)
        team1 = team_lookup.get(team1_id)
        team2 = team_lookup.get(team2_id)
        if not winner or not team1 or not team2:
            continue

        team1_name = team1.team_name
        team2_name = team2.team_name
        alias_forms = {
            f"{team1_name}/{team2_name}",
            f"{team2_name}/{team1_name}",
            f"{team1_name} / {team2_name}",
            f"{team2_name} / {team1_name}",
            f"{team1_name} {team2_name}",
            f"{team2_name} {team1_name}",
            f"{team1_name}/{team2_name} Winner",
            f"{team2_name}/{team1_name} Winner",
            f"Winner of {team1_name}/{team2_name}",
            f"Winner of {team2_name}/{team1_name}",
            f"Winner of {team1_name} vs {team2_name}",
            f"Winner of {team2_name} vs {team1_name}",
        }
        for alias in alias_forms:
            resolution_map[normalize_team_key(alias)] = winner

    return resolution_map


def apply_resolved_placeholder_picks(
    entries_df: pd.DataFrame,
    placeholder_resolution_map: Dict[str, str],
) -> Tuple[pd.DataFrame, List[Tuple[str, str, str]]]:
    if not placeholder_resolution_map:
        return entries_df, []

    resolved_entries = entries_df.copy()
    replacements: List[Tuple[str, str, str]] = []

    for idx, row in resolved_entries.iterrows():
        entry_name = str(row.get("name") or "").strip()
        for column_idx in range(1, 11):
            column = f"team_{column_idx}"
            raw_value = row.get(column)
            if pd.isna(raw_value):
                continue
            normalized = normalize_team_key(raw_value)
            replacement = placeholder_resolution_map.get(normalized)
            if replacement and str(raw_value).strip() != replacement:
                resolved_entries.at[idx, column] = replacement
                replacements.append((entry_name, str(raw_value).strip(), replacement))

    return resolved_entries, replacements


def build_dashboard_payload(
    leaderboard: pd.DataFrame,
    *,
    updated_at: str,
    current_completed_games: int,
    trend_reference_completed_games: int,
    trend_baseline_updated: bool,
    latest_completed_game: Optional[str],
    snapshot_state: Dict[str, object],
    team_status_map: Dict[str, Dict[str, object]],
    current_round_code: Optional[int],
) -> Dict[str, object]:
    round_labels = {
        64: "Round of 64",
        32: "Round of 32",
        16: "Sweet 16",
        8: "Elite 8",
        4: "Final Four",
        2: "Championship",
    }
    top_by_odds = leaderboard.sort_values(["live_win_rate", "current_score"], ascending=[False, False]).reset_index(drop=True)
    biggest_riser = leaderboard.sort_values(["trend_rank_delta", "trend_odds_delta"], ascending=[False, False]).iloc[0]
    rows = []
    for row in leaderboard.itertuples(index=False):
        team_badges = []
        for team_name in [team.strip() for team in str(row.teams).split(",") if team.strip()]:
            team_state = team_status_map.get(team_name, {"current_round_status": "not_played_current_round", "alive": True})
            team_badges.append(
                {
                    "team": team_name,
                    "alive": bool(team_state.get("alive", True)),
                    "current_round_status": str(team_state.get("current_round_status") or "not_played_current_round"),
                }
            )
        rows.append(
            {
                "name": row.name,
                "current_rank": int(row.current_rank),
                "current_score": int(row.current_score),
                "live_win_rate": float(row.live_win_rate),
                "top4_rate": float(row.top4_rate),
                "remaining_live_teams": int(row.remaining_live_teams),
                "p50_score": float(row.p50_score),
                "best_case_score": int(row.best_case_score),
                "top3_rate": float(row.top3_rate),
                "top10_rate": float(row.top10_rate),
                "last_rate": float(row.last_rate),
                "trend_rank_delta": float(row.trend_rank_delta),
                "trend_score_delta": float(row.trend_score_delta),
                "trend_odds_delta": float(row.trend_odds_delta),
                "trend_rank_direction": str(row.trend_rank_direction),
                "current_rank_label": int(row.current_rank),
                "teams": team_badges,
            }
        )

    return {
        "updated_at": updated_at,
        "headline": {
            "title": "March Madness Pool Live Leaderboard",
            "subtitle": "Current standings and odds to win from the latest completed-game state.",
        },
        "trend": {
            "baseline_updated_this_run": bool(trend_baseline_updated),
            "current_completed_games": int(current_completed_games),
            "trend_reference_completed_games": int(trend_reference_completed_games),
            "latest_completed_game": latest_completed_game or "",
            "current_round_label": round_labels.get(current_round_code, ""),
        },
        "summary_cards": [
            {
                "label": "Current Leader",
                "value": str(leaderboard.iloc[0]["name"]),
                "detail": f"{int(leaderboard.iloc[0]['current_score'])} points",
            },
            {
                "label": "Odds Favorite",
                "value": str(top_by_odds.iloc[0]["name"]),
                "detail": f"{top_by_odds.iloc[0]['live_win_rate'] * 100:.1f}% win odds",
            },
            {
                "label": "Biggest Riser",
                "value": str(biggest_riser["name"]),
                "detail": (
                    "No movement since the baseline"
                    if float(biggest_riser["trend_rank_delta"]) == 0
                    else f"{abs(int(biggest_riser['trend_rank_delta']))} rank spots"
                ),
            },
        ],
        "snapshot_state": snapshot_state,
        "rows": rows,
    }


def build_stock_label(movement: float) -> str:
    if movement > 0.5:
        return "▲ Up"
    if movement < -0.5:
        return "▼ Down"
    return "— Flat"


def build_polished_leaderboard(leaderboard: pd.DataFrame) -> pd.DataFrame:
    polished = leaderboard.copy()
    polished["Rank"] = polished["current_score"].rank(method="min", ascending=False).astype(int)
    polished["Teams Left"] = polished["remaining_live_teams"]
    polished["Chance to Win"] = polished["live_win_rate"]
    polished["Live Projection"] = polished["p50_score"]
    polished["Max"] = polished["best_case_score"]
    polished["Top 3%"] = polished["top3_rate"]
    polished["Top 10%"] = polished["top10_rate"]
    polished["Last %"] = polished["last_rate"]
    polished = polished.sort_values(["Rank", "current_score"], ascending=[True, False])
    polished["Rank"] = range(1, len(polished) + 1)
    polished["Name"] = polished["name"]
    polished["Current Score"] = polished["current_score"]
    return polished[
        [
            "Rank",
            "Name",
            "Current Score",
            "Teams Left",
            "Chance to Win",
            "Live Projection",
            "Max",
            "Top 3%",
            "Top 10%",
            "Last %",
        ]
    ]


def build_game_state_df(live_state: Dict[str, LiveNodeState]) -> pd.DataFrame:
    round_labels = {
        68: "First Four",
        64: "Round of 64",
        32: "Round of 32",
        16: "Sweet 16",
        8: "Elite 8",
        4: "Final Four",
        2: "Championship",
    }
    rows = []
    for node_id, state in sorted(live_state.items(), key=lambda item: (ROUND_SORT_PRIORITY.get(item[1].game.round_code, 99), item[1].game.region, item[1].game.slot)):
        if state.game.round_code == 68:
            continue
        participants = state.matched_participants or (None, None)
        if not participants[0] or not participants[1]:
            continue
        prob_a = ""
        prob_b = ""
        if state.market_probs and participants[0] in state.market_probs and participants[1] in state.market_probs:
            prob_a = state.market_probs[participants[0]]
            prob_b = state.market_probs[participants[1]]
        rows.append(
            {
                "Round": round_labels.get(state.game.round_code, str(state.game.round_code)),
                "Region": state.game.region,
                "Matchup": f"{participants[0]} vs {participants[1]}",
                "Status": state.score_status or "scheduled",
                "Winner": state.fixed_winner,
                "Has Market Odds": bool(state.market_probs),
                "Prob A": prob_a,
                "Prob B": prob_b,
                "node_id": node_id,
            }
        )
    return pd.DataFrame(rows)


def write_excel(output_path: Path, leaderboard: pd.DataFrame, status_df: pd.DataFrame, game_state_df: pd.DataFrame) -> None:
    polished_leaderboard = build_polished_leaderboard(leaderboard)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        status_df.to_excel(writer, sheet_name="Status", index=False)
        game_state_df.to_excel(writer, sheet_name="Game State", index=False)
        polished_leaderboard.to_excel(
            writer,
            sheet_name="Leaderboard",
            index=False,
        )
        workbook = writer.book
        leaderboard_ws = writer.sheets["Leaderboard"]
        game_state_ws = writer.sheets["Game State"]
        status_ws = writer.sheets["Status"]

        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        alt_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        def style_sheet(ws, widths, percent_cols=None, integer_cols=None, decimal_cols=None):
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            ws.row_dimensions[1].height = 20
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for row_idx in range(2, ws.max_row + 1):
                if row_idx % 2 == 0:
                    for cell in ws[row_idx]:
                        cell.fill = alt_fill
            for column_letter, width in widths.items():
                ws.column_dimensions[column_letter].width = width
            percent_cols = percent_cols or []
            integer_cols = integer_cols or []
            decimal_cols = decimal_cols or []
            for column_letter in percent_cols:
                for cell in ws[column_letter][1:]:
                    cell.number_format = "0.0%"
                    cell.alignment = right_alignment
            for column_letter in integer_cols:
                for cell in ws[column_letter][1:]:
                    cell.number_format = "0"
                    cell.alignment = right_alignment
            for column_letter in decimal_cols:
                for cell in ws[column_letter][1:]:
                    cell.number_format = "0.0"
                    cell.alignment = right_alignment

        def header_to_column_map(ws):
            return {cell.value: cell.column_letter for cell in ws[1] if cell.value}

        leaderboard_cols = header_to_column_map(leaderboard_ws)
        style_sheet(
            leaderboard_ws,
            widths={"A": 8, "B": 24, "C": 14, "D": 12, "E": 14, "F": 14, "G": 12, "H": 12, "I": 12, "J": 12},
            percent_cols=[
                leaderboard_cols["Chance to Win"],
                leaderboard_cols["Top 3%"],
                leaderboard_cols["Top 10%"],
                leaderboard_cols["Last %"],
            ],
            integer_cols=[
                leaderboard_cols["Rank"],
                leaderboard_cols["Current Score"],
                leaderboard_cols["Teams Left"],
                leaderboard_cols["Live Projection"],
                leaderboard_cols["Max"],
            ],
        )
        for cell in leaderboard_ws[leaderboard_cols["Name"]][1:]:
            cell.alignment = left_alignment

        style_sheet(
            game_state_ws,
            widths={"A": 16, "B": 14, "C": 32, "D": 14, "E": 18, "F": 16, "G": 10, "H": 10, "I": 16},
            percent_cols=["G", "H"],
        )
        for col in ["A", "B", "C", "D", "E", "I"]:
            for cell in game_state_ws[col][1:]:
                cell.alignment = left_alignment
        for cell in game_state_ws["F"][1:]:
            cell.alignment = center_alignment

        status_ws.freeze_panes = "A2"
        status_ws.auto_filter.ref = status_ws.dimensions
        status_ws.column_dimensions["A"].width = 34
        status_ws.column_dimensions["B"].width = 80
        for cell in status_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        for cell in status_ws["A"][1:] + status_ws["B"][1:]:
            cell.alignment = left_alignment


def sync_site_outputs(base_dir: Path) -> List[Path]:
    site_dir = base_dir / SITE_OUTPUT_DIR
    site_dir.mkdir(parents=True, exist_ok=True)
    synced_paths: List[Path] = []

    for filename in (OUTPUT_CSV, OUTPUT_PREVIOUS_CSV, OUTPUT_DASHBOARD_JSON, SNAPSHOT_STATE_FILE):
        source = base_dir / filename
        if not source.exists():
            continue
        destination = site_dir / filename
        shutil.copyfile(source, destination)
        synced_paths.append(destination)

    return synced_paths


def main() -> None:
    base_dir = Path.cwd()
    load_env_file(base_dir / ".env")

    workbook_path = base_dir / TOURNAMENT_FILE
    if not workbook_path.exists():
        raise FileNotFoundError(f"Could not find tournament workbook: {workbook_path}")

    entries_path = resolve_pool_entries_path(base_dir)
    raw_entries_df = read_pool_entries(entries_path)

    team_lookup, _games, round_multipliers, logistic_k, rating_col = load_model_inputs(workbook_path)
    first_four_lookup = load_first_four_games(workbook_path)
    team_lookup_by_name = {team.team_name: team for team in team_lookup.values()}
    team_name_index = build_team_name_index(team_lookup)
    team_id_index = build_team_id_index(team_lookup)

    resolved_winners_path = base_dir / RESOLVED_WINNERS_FILE
    saved_winners = load_resolved_winners(resolved_winners_path)
    nodes = build_bracket_template(workbook_path, saved_winners)
    scoreboard_dates = _iter_tournament_scoreboard_dates()
    scoreboard_urls = [
        "https://ncaa-api.henrygd.me/scoreboard/"
        f"basketball-men/d1/{dt:%Y/%m/%d}/all-conf"
        for dt in scoreboard_dates
    ]
    try:
        score_events, score_credits_remaining = load_ncaa_scoreboard_games()
    except urllib.error.HTTPError as exc:
        if exc.code != 403:
            raise
        print("NCAA scoreboard fetch failed with HTTP 403; falling back to saved state only.")
        score_events = []
        score_credits_remaining = None
    except urllib.error.URLError:
        print("NCAA scoreboard fetch failed; falling back to saved state only.")
        score_events = []
        score_credits_remaining = None
    try:
        odds_events, odds_credits_remaining = load_the_odds_api_odds(score_events)
    except urllib.error.HTTPError as exc:
        if exc.code != 403:
            raise
        print("Odds API fetch failed with HTTP 403; using fallback model for all remaining games.")
        odds_events = []
        odds_credits_remaining = None
    except urllib.error.URLError:
        print("Odds API fetch failed; using fallback model for all remaining games.")
        odds_events = []
        odds_credits_remaining = None
    odds_by_event_id, odds_by_pair, odds_before_count, odds_after_match_count = build_market_probabilities(
        odds_events,
        team_name_index,
        team_id_index,
    )
    live_state, fixed_winners, counts = build_live_state(
        nodes,
        score_events,
        odds_by_event_id,
        odds_by_pair,
        team_name_index,
        team_id_index,
        saved_winners,
    )
    merged_winners = dict(saved_winners)
    merged_winners.update(fixed_winners)
    save_resolved_winners(resolved_winners_path, merged_winners)
    placeholder_resolution_map = build_first_four_pick_resolution_map(
        first_four_lookup,
        team_lookup,
        merged_winners,
    )
    entries_df, placeholder_replacements = apply_resolved_placeholder_picks(
        raw_entries_df,
        placeholder_resolution_map,
    )
    entries_df, entries_matrix, ordered_team_names = build_entries_matrix(entries_df, team_lookup)
    team_index = {team_name: idx for idx, team_name in enumerate(ordered_team_names)}
    new_winner_count = sum(1 for key, value in fixed_winners.items() if saved_winners.get(key) != value)
    current_snapshot_key = build_snapshot_key(merged_winners)
    newly_completed_games = describe_completed_games(nodes, live_state, saved_winners, merged_winners)
    latest_completed_game = newly_completed_games[-1] if newly_completed_games else ""
    credits_remaining = odds_credits_remaining or score_credits_remaining
    print(f"Odds events before filtering: {odds_before_count}")
    print(f"Odds events after filtering: {odds_after_match_count}")

    print("\nRound of 64 scoreboard mapping debug")
    print("===================================")
    for node_id, state in sorted(
        ((node_id, state) for node_id, state in live_state.items() if state.game.round_code == 64),
        key=lambda item: (item[1].game.region, item[1].game.slot),
    ):
        participants = state.matched_participants or (None, None)
        print(f"node_id: {node_id}")
        print(f"resolved matchup teams: {participants[0]} vs {participants[1]}")
        print(f"score_status: {state.score_status}")
        print(f"fixed_winner: {state.fixed_winner}")
        print(f"matched_to_scoreboard_event: {bool(state.score_event_id)}")

    matched_score_event_ids = {
        str(state.score_event_id)
        for state in live_state.values()
        if state.score_event_id
    }
    print("\nUnmatched NCAA scoreboard events")
    print("===============================")
    unmatched_score_events_found = False
    for event in score_events:
        event_id = str(event.get("id") or "")
        if event_id in matched_score_event_ids:
            continue
        unmatched_score_events_found = True
        scores = event.get("scores", [])
        if isinstance(scores, list) and len(scores) >= 2:
            raw_names = [str(item.get("name") or "") for item in scores[:2]]
        else:
            raw_names = [str(event.get("away_team") or ""), str(event.get("home_team") or "")]
        print(f"event_id: {event_id}")
        print(f"raw teams: {raw_names[0]} vs {raw_names[1]}")
        print(f"completed: {event.get('completed')}")
        print(f"status: {event.get('status')}")
    if not unmatched_score_events_found:
        print("None")

    first_four_report, unresolved_placeholders = validate_first_four_placeholders(
        nodes,
        fixed_winners,
        team_name_index,
        team_id_index,
    )
    print(first_four_report)
    if unresolved_placeholders:
        details = []
        for placeholder_id in unresolved_placeholders:
            node = nodes[placeholder_id]
            team_a = canonicalize_name(node.source_a, team_name_index, team_id_index)
            team_b = canonicalize_name(node.source_b, team_name_index, team_id_index)
            available = [team for team in [team_a, team_b] if team]
            details.append(
                f"{placeholder_id}; expected source matchup: {node.source_a} vs {node.source_b}; "
                f"available candidate teams: {', '.join(available) if available else 'none'}"
            )
        raise ValueError("Unresolved First Four placeholders:\n" + "\n".join(details))

    unresolved_names = find_unresolved_bracket_names(nodes, fixed_winners, team_name_index, team_id_index)
    print("Unresolved bracket names before simulation:")
    if unresolved_names:
        for name in unresolved_names:
            print(f" - {name}")
        raise ValueError(
            "Unresolved team names found in current bracket state: "
            + ", ".join(unresolved_names)
        )
    print(" - None")
    if placeholder_replacements:
        print("Resolved placeholder entry picks")
        print("==============================")
        for entry_name, original_value, resolved_value in placeholder_replacements:
            print(f" - {entry_name}: {original_value} -> {resolved_value}")

    current_team_scores, eliminated_teams = compute_current_state(
        nodes,
        live_state,
        team_lookup_by_name,
        round_multipliers,
    )
    current_team_score_vector = np.zeros(len(ordered_team_names), dtype=np.int16)
    for team_name, score in current_team_scores.items():
        current_team_score_vector[team_index[team_name]] = score
    current_entry_scores = entries_matrix @ current_team_score_vector

    print(build_validation_summary(counts, credits_remaining))
    zero_counts_diagnostic = build_zero_counts_diagnostic(
        odds_events,
        score_events,
        counts,
        team_name_index,
    )
    if zero_counts_diagnostic:
        print(zero_counts_diagnostic)

    simulation_count = get_simulation_count()
    rng = np.random.default_rng(20260319)
    sim_score_matrix = np.zeros((simulation_count, len(entries_df)), dtype=np.int16)
    finish_matrix = np.zeros((simulation_count, len(entries_df)), dtype=float)
    win_share_matrix = np.zeros((simulation_count, len(entries_df)), dtype=float)
    top3_share_matrix = np.zeros((simulation_count, len(entries_df)), dtype=float)
    top4_share_matrix = np.zeros((simulation_count, len(entries_df)), dtype=float)
    top10_share_matrix = np.zeros((simulation_count, len(entries_df)), dtype=float)
    last_share_matrix = np.zeros((simulation_count, len(entries_df)), dtype=float)

    for sim_idx in range(simulation_count):
        additional_team_scores = simulate_remaining_bracket(
            nodes=nodes,
            live_state=live_state,
            fixed_winners=fixed_winners,
            team_lookup_by_name=team_lookup_by_name,
            team_name_index=team_name_index,
            team_id_index=team_id_index,
            round_multipliers=round_multipliers,
            logistic_k=logistic_k,
            rng=rng,
        )
        team_score_vector = current_team_score_vector.copy()
        for team_name, score in additional_team_scores.items():
            team_score_vector[team_index[team_name]] += score
        total_scores = entries_matrix @ team_score_vector
        sim_score_matrix[sim_idx] = total_scores - current_entry_scores
        finish_values, top_shares, last_place_share = calculate_finish_metrics(total_scores)
        finish_matrix[sim_idx] = finish_values
        win_share_matrix[sim_idx] = top_shares[1]
        top3_share_matrix[sim_idx] = top_shares[3]
        top4_share_matrix[sim_idx] = top_shares[4]
        top10_share_matrix[sim_idx] = top_shares[10]
        last_share_matrix[sim_idx] = last_place_share

    summary = summarize_results(
        entries_df=entries_df,
        current_scores=current_entry_scores,
        sim_score_matrix=sim_score_matrix,
        finish_matrix=finish_matrix,
        win_share_matrix=win_share_matrix,
        top3_share_matrix=top3_share_matrix,
        top4_share_matrix=top4_share_matrix,
        top10_share_matrix=top10_share_matrix,
        last_share_matrix=last_share_matrix,
        eliminated_teams=eliminated_teams,
    )
    leaderboard = summary.sort_values(["current_score", "live_win_rate"], ascending=[False, False]).reset_index(drop=True)
    snapshot_state_path = base_dir / SNAPSHOT_STATE_FILE
    existing_snapshot_state = load_snapshot_state(snapshot_state_path)
    output_csv_path = base_dir / OUTPUT_CSV
    previous_csv_path = base_dir / OUTPUT_PREVIOUS_CSV
    baseline_updated_this_run = False
    previous_snapshot_key = str(existing_snapshot_state.get("previous_snapshot_key") or "")
    if output_csv_path.exists():
        prior_current_snapshot_key = str(existing_snapshot_state.get("current_snapshot_key") or "")
        if prior_current_snapshot_key and prior_current_snapshot_key != current_snapshot_key:
            shutil.copyfile(output_csv_path, previous_csv_path)
            previous_snapshot_key = prior_current_snapshot_key
            baseline_updated_this_run = True
        elif not previous_csv_path.exists():
            shutil.copyfile(output_csv_path, previous_csv_path)
            previous_snapshot_key = prior_current_snapshot_key or current_snapshot_key

    baseline_df = pd.read_csv(previous_csv_path) if previous_csv_path.exists() else None
    if baseline_df is not None and not previous_snapshot_key:
        previous_snapshot_key = str(existing_snapshot_state.get("current_snapshot_key") or current_snapshot_key)
    leaderboard = apply_trend_baseline(leaderboard, baseline_df)
    output_timestamp = datetime.now(timezone.utc).isoformat()
    previous_completed_games = int(existing_snapshot_state.get("previous_completed_games") or counts["completed_games"])
    if baseline_updated_this_run:
        previous_completed_games = int(existing_snapshot_state.get("current_completed_games") or counts["completed_games"])
    elif baseline_df is None:
        previous_completed_games = counts["completed_games"]
    snapshot_state = {
        "current_snapshot_key": current_snapshot_key,
        "previous_snapshot_key": previous_snapshot_key,
        "current_completed_games": int(counts["completed_games"]),
        "previous_completed_games": int(previous_completed_games),
        "trend_baseline_updated_this_run": baseline_updated_this_run,
        "newly_completed_games_count": int(new_winner_count),
        "newly_completed_games": newly_completed_games,
        "latest_completed_game": latest_completed_game,
        "updated_at": output_timestamp,
    }
    status_rows = [
        ("ncaa_scoreboard_base_date_used", get_scoreboard_base_date().isoformat()),
        ("ncaa_scoreboard_days_used", str(get_scoreboard_days())),
        ("scoreboard_date_urls_fetched", "\n".join(scoreboard_urls)),
        ("resolved_winners_loaded_from_disk", str(len(saved_winners))),
        ("new_winners_added_this_run", str(new_winner_count)),
        ("total_resolved_winners_after_saving", str(len(merged_winners))),
        ("current_snapshot_key", current_snapshot_key),
        ("previous_snapshot_key", previous_snapshot_key or current_snapshot_key),
        ("trend_baseline_updated_this_run", str(baseline_updated_this_run)),
        ("latest_completed_game", latest_completed_game or "None"),
        ("completed_games", str(counts["completed_games"])),
        ("live_games", str(counts["live_games"])),
        ("remaining_games_with_odds", str(counts["remaining_games_with_odds"])),
        ("remaining_games_using_fallback_model", str(counts["remaining_games_fallback"])),
        ("simulations_run", str(simulation_count)),
        ("win_probability_sum", f"{leaderboard['live_win_rate'].sum():.6f}"),
        ("top3_share_sum", f"{leaderboard['top3_rate'].sum():.6f}"),
        ("last_place_share_sum", f"{leaderboard['last_rate'].sum():.6f}"),
        ("output_timestamp", output_timestamp),
    ]
    status_df = pd.DataFrame(status_rows, columns=["metric", "value"])
    game_state_df = build_game_state_df(live_state)
    team_status_map, current_round_code = build_team_status_map(nodes, live_state)
    print("\nActive Round of 64 debug")
    print("========================")
    for node_id in ["R64_W_05", "R64_S_01", "R64_M_01", "R64_M_05"]:
        state = live_state[node_id]
        participants = state.matched_participants or (None, None)
        print(f"node_id: {node_id}")
        print(f"source_a: {nodes[node_id].source_a}")
        print(f"source_b: {nodes[node_id].source_b}")
        print(f"resolved team_a: {participants[0]}")
        print(f"resolved team_b: {participants[1]}")
    leaderboard.to_csv(output_csv_path, index=False)
    write_excel(base_dir / OUTPUT_XLSX, leaderboard, status_df, game_state_df)
    dashboard_payload = build_dashboard_payload(
        leaderboard,
        updated_at=output_timestamp,
        current_completed_games=counts["completed_games"],
        trend_reference_completed_games=previous_completed_games,
        trend_baseline_updated=baseline_updated_this_run,
        latest_completed_game=latest_completed_game,
        snapshot_state=snapshot_state,
        team_status_map=team_status_map,
        current_round_code=current_round_code,
    )
    (base_dir / OUTPUT_DASHBOARD_JSON).write_text(json.dumps(dashboard_payload, indent=2), encoding="utf-8")
    save_snapshot_state(snapshot_state_path, snapshot_state)
    synced_site_paths = sync_site_outputs(base_dir)

    print("\nRun status")
    print("==========")
    for metric, value in status_rows:
        print(f"{metric}: {value}")

    print(f"\nUsed team rating column: {rating_col}")
    print(f"Used logistic_k: {logistic_k}")
    print(f"Simulations run: {simulation_count}")
    print(f"Read pool entries from: {entries_path.name}")
    print(f"Wrote CSV output: {OUTPUT_CSV}")
    print(f"Wrote Excel output: {OUTPUT_XLSX}")
    if synced_site_paths:
        for path in synced_site_paths:
            print(f"Synced site output: {path.relative_to(base_dir)}")


if __name__ == "__main__":
    main()
